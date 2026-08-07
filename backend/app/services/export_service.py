"""
Export service — 확정 시간표를 학부 배포 양식 엑셀로 출력합니다.

서식(폰트·크기·색·열너비·행높이·블록 배치)은 학부 원본 파일에서 실측한 값이며,
``test/verify_export_format.py``가 원본과 셀 단위로 대조합니다.
"""

from __future__ import annotations

import io
import logging
import os
from typing import Any

from sqlalchemy.orm import Session

from app.models.course import Course
from app.models.professor import Professor
from app.models.room import Room
from app.models.timetable import Assignment, Timetable
from app.services.timetable_service import get_timetable

from app.algorithm.constraints import BLOCKED_SLOTS, DAYS, PERIODS

logger = logging.getLogger(__name__)

DAY_LABELS = {"MON": "월", "TUE": "화", "WED": "수", "THU": "목", "FRI": "금"}

# ── Workbook layout, reverse-engineered from the department's real file ──────
# Every sheet is tiled with "blocks": an anchor cell holding the block title,
# 월~금 to its right, and one row per period below it.
BLOCK_COL_STRIDE = 7    # B → I → P …  (라벨 1열 + 요일 5열 + 간격 1열)
BLOCKS_PER_ROW = 3      # 이론실/실습실/학과별. 교수님별은 5개.
PROF_BLOCKS_PER_ROW = 5
FIRST_ROW = 2           # 1행은 시트 제목
FIRST_COL = 2           # column B

# ── 원본 파일에서 실측한 서식값 ────────────────────────────────────────────
# (블록 행 stride = 교시 수 + 2  →  9교시면 11, 10교시면 12)
FONT_NAME = "맑은 고딕"
GRID_RGB = "2E75B6"     # 원본 테마 accent1(5B9BD5)에 tint -0.25 적용한 실제 색

SZ_SHEET_TITLE = 16     # B1 시트 제목
SZ_BLOCK_TITLE = 10     # 강의실/교수 이름
SZ_HEADER = 8           # 요일 헤더 · 교시 라벨
SZ_COURSE = 10          # 수업 셀
SZ_EMPTY = 11           # 빈 칸 · 비교과

W_SPACER_FIRST = 2.5    # A열
W_SPACER = 2.125        # 블록 사이 간격 열
W_BLOCK_COL = 11.625    # 라벨/요일 열
H_TITLE_ROW = 30.0
H_BLOCK_ROW = 32.1
H_SPACER_ROW = 17.1

# 0교시는 온라인 수업 전용이라 강의실을 점유하지 않는다.
ONLINE_PERIOD = 0

BLOCKED_LABEL = "비교과"

# 강의실 시트에는 온라인 전용인 0교시 행을 두지 않는다.
ROOM_PERIODS = [p for p in PERIODS if p != ONLINE_PERIOD]


def _build_grid(db: Session, timetable_id: int) -> tuple[list[dict], dict]:
    """Build a (assignments, lookup) tuple for rendering."""
    assignments = db.query(Assignment).filter(Assignment.timetable_id == timetable_id).all()
    course_map = {c.id: c for c in db.query(Course).all()}
    room_map = {r.id: r for r in db.query(Room).all()}
    prof_map = {p.id: p for p in db.query(Professor).all()}

    rows = []
    for a in assignments:
        c = course_map.get(a.course_id)
        r = room_map.get(a.room_id)
        p = prof_map.get(c.professor_id) if c else None
        rows.append(
            {
                "day": a.day,
                "day_label": DAY_LABELS.get(a.day, a.day),
                "period": a.start_period,
                "course": c.course_name if c else "?",
                # 온라인 수업(0교시)은 강의실이 없다.
                "room": r.room_name if r else ("온라인" if a.room_id is None else "?"),
                "professor": p.name if p else "?",
                "department": c.department if c else "?",
                "grade": c.target_grade if c else "?",
            }
        )
    return rows, {}


def _collect(db: Session, timetable_id: int) -> list[dict]:
    """Flatten a timetable's assignments into render-ready records."""
    assignments = db.query(Assignment).filter(Assignment.timetable_id == timetable_id).all()
    course_map = {c.id: c for c in db.query(Course).all()}
    room_map = {r.id: r for r in db.query(Room).all()}
    prof_map = {p.id: p for p in db.query(Professor).all()}

    out = []
    for a in assignments:
        c = course_map.get(a.course_id)
        if not c:
            continue
        r = room_map.get(a.room_id)
        p = prof_map.get(c.professor_id)
        out.append(
            {
                "day": a.day,
                "period": a.start_period,
                "course": c.course_name,
                "section": (c.class_section or "").strip(),
                "professor": p.name if p else "",
                "prof_id": c.professor_id,
                "prof_no": p.employee_number if p else "",
                "room": r.room_name if r else "",
                "room_id": a.room_id,
                "is_computer_room": bool(r.is_computer_room) if r else False,
                "department": c.department,
                "grade": c.target_grade,
                "course_id": c.id,
            }
        )
    return out


def _runs(records: list[dict], key_fn) -> list[tuple[dict, int, int]]:
    """
    Collapse consecutive periods of the same course/day into a single run so it
    can be rendered as one merged cell, exactly like the department's file.

    Returns (record, start_period, length) tuples.
    """
    buckets: dict[Any, list[dict]] = {}
    for rec in records:
        buckets.setdefault((key_fn(rec), rec["day"], rec["course_id"]), []).append(rec)

    out = []
    for group in buckets.values():
        group.sort(key=lambda z: z["period"])
        start = prev = None
        for rec in group:
            if start is None:
                start = prev = rec
                length = 1
            elif rec["period"] == prev["period"] + 1:
                prev = rec
                length += 1
            else:
                out.append((start, start["period"], length))
                start = prev = rec
                length = 1
        if start is not None:
            out.append((start, start["period"], length))
    return out


def _period_label(period: int, style: str) -> str:
    hour = 8 + period
    if style == "prof":
        return f"{period}교시    {hour}:00"
    return f"{period} ({hour:02d}:00)"


def _make_styles():
    """Style bundle matching the department's original workbook exactly."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    side = Side(style="thin")
    border = Border(left=side, right=side, top=side, bottom=side)
    fill = PatternFill("solid", fgColor=GRID_RGB)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    return {
        "border": border,
        "fill": fill,
        "center": center,
        "sheet_title": Font(name=FONT_NAME, size=SZ_SHEET_TITLE, bold=True),
        "block_title": Font(name=FONT_NAME, size=SZ_BLOCK_TITLE, bold=True),
        "header": Font(name=FONT_NAME, size=SZ_HEADER, bold=True),
        "course": Font(name=FONT_NAME, size=SZ_COURSE, bold=False),
        "empty": Font(name=FONT_NAME, size=SZ_EMPTY, bold=False),
        "blocked": Font(name=FONT_NAME, size=SZ_EMPTY, bold=True),
    }


def _draw_block(ws, top: int, left: int, title: str, runs, label_style: str,
                st, periods: list[int]) -> None:
    """
    Draw one block: title + 월~금 header + one row per period.

    Everything except a filled-in class cell carries the grid fill, exactly as
    in the original workbook (classes read as white cells on a blue grid).
    """
    tc = ws.cell(row=top, column=left, value=title)
    tc.font, tc.fill, tc.alignment, tc.border = st["block_title"], st["fill"], st["center"], st["border"]

    for di, day in enumerate(DAYS):
        h = ws.cell(row=top, column=left + 1 + di, value=DAY_LABELS[day])
        h.font, h.fill, h.alignment, h.border = st["header"], st["fill"], st["center"], st["border"]

    # 학과별 블록은 코호트가 다르면 같은 슬롯에 두 과목이 올 수 있다.
    by_slot: dict[tuple[str, int], list] = {}
    for rec, start, length in runs:
        by_slot.setdefault((rec["day"], start), []).append((rec, length))

    occupied: set[tuple[str, int]] = set()
    for rec, start, length in runs:
        for k in range(1, length):
            occupied.add((rec["day"], start + k))

    for pi, period in enumerate(periods):
        row = top + 1 + pi
        lc = ws.cell(row=row, column=left, value=_period_label(period, label_style))
        lc.font, lc.fill, lc.alignment, lc.border = st["header"], st["fill"], st["center"], st["border"]

        for di, day in enumerate(DAYS):
            col = left + 1 + di
            cell = ws.cell(row=row, column=col)
            cell.alignment = st["center"]
            cell.border = st["border"]

            if (day, period) in occupied:
                cell.font, cell.fill = st["course"], st["fill"]
                continue

            hits = by_slot.get((day, period))
            if hits:
                cell.value = "\n".join(rec["text"] for rec, _ in hits)
                cell.font = st["course"]          # 수업 칸은 채우기 없음(흰색)
                length = max(n for _, n in hits)
                if length > 1:
                    ws.merge_cells(start_row=row, start_column=col,
                                   end_row=row + length - 1, end_column=col)
                    for k in range(length):
                        m = ws.cell(row=row + k, column=col)
                        m.border, m.font = st["border"], st["course"]
            elif (day, period) in BLOCKED_SLOTS:
                cell.value = BLOCKED_LABEL
                cell.font, cell.fill = st["blocked"], st["fill"]
            else:
                cell.font, cell.fill = st["empty"], st["fill"]


def _add_sheet(wb, name: str, sheet_title: str, blocks: list[tuple[str, list]],
               label_style: str, st, periods: list[int], per_row: int):
    """Lay blocks out on a new sheet with the original's spacing and sizing."""
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(title=name[:31])
    ws.page_setup.orientation = "landscape"
    ws.sheet_view.showGridLines = False

    row_stride = len(periods) + 2          # 제목행 + 교시행들 + 간격행

    if sheet_title:
        t = ws.cell(row=1, column=FIRST_COL, value=sheet_title)
        t.font = st["sheet_title"]
        t.alignment = st["center"]
    ws.row_dimensions[1].height = H_TITLE_ROW

    ws.column_dimensions["A"].width = W_SPACER_FIRST

    for i, (title, runs) in enumerate(blocks):
        top = FIRST_ROW + (i // per_row) * row_stride
        left = FIRST_COL + (i % per_row) * BLOCK_COL_STRIDE
        _draw_block(ws, top, left, title, runs, label_style, st, periods)

        for r in range(top, top + len(periods) + 1):
            ws.row_dimensions[r].height = H_BLOCK_ROW
        spacer_row = top + len(periods) + 1
        ws.row_dimensions[spacer_row].height = H_SPACER_ROW

        for c in range(left, left + 6):
            ws.column_dimensions[get_column_letter(c)].width = W_BLOCK_COL
        ws.column_dimensions[get_column_letter(left + 6)].width = W_SPACER

    return ws


def export_excel(db: Session, timetable_id: int) -> bytes:
    """
    Generate the department-format workbook: one sheet per view (강의실 / 교수 /
    학과별), each tiled with per-room / per-professor / per-grade block grids.

    Sheets are derived from the data, never hardcoded, so a semester with
    different departments, rooms or professors exports correctly with no
    code change.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl 라이브러리가 설치되어 있지 않습니다.")

    records = _collect(db, timetable_id)

    st = _make_styles()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def with_text(recs, fmt):
        for r in recs:
            r = dict(r)
            r["text"] = fmt(r)
            yield r

    def joined(*parts) -> str:
        return "_".join(str(p) for p in parts if p)

    # ── 1) 강의실 기준: 이론실 / 실습실 ─────────────────────────────────────
    # 0교시는 온라인 전용이라 강의실 시트에서는 제외한다.
    rooms = sorted(
        {(r["room_id"], r["room"], r["is_computer_room"]) for r in records if r["room"]},
        key=lambda z: str(z[1]),
    )
    room_caps = {r.id: r.capacity for r in db.query(Room).all()}
    for is_pc, sheet_name, sheet_title in (
        (False, "이론실", "이론실"), (True, "실습실", "실습실(PC)")
    ):
        blocks = []
        for rid, rname, pc in rooms:
            if pc != is_pc:
                continue
            recs = [
                r for r in records
                if r["room_id"] == rid and r["period"] != ONLINE_PERIOD
            ]
            cap = room_caps.get(rid)
            title = f"{rname}_{cap}석" if cap else str(rname)
            blocks.append(
                (title, _runs(list(with_text(recs, lambda r: joined(
                    r["course"], r["section"], r["professor"]))), lambda r: r["room_id"]))
            )
        if blocks:
            _add_sheet(wb, sheet_name, sheet_title, blocks, "room", st,
                       ROOM_PERIODS, BLOCKS_PER_ROW)

    # ── 2) 교수 기준 ────────────────────────────────────────────────────────
    profs = sorted(
        {(r["prof_id"], r["professor"], r["prof_no"]) for r in records if r["prof_id"]},
        key=lambda z: str(z[1]),
    )
    blocks = []
    for pid, pname, pno in profs:
        recs = [r for r in records if r["prof_id"] == pid]
        title = f"{pname}\n({pno})" if pno else str(pname)
        blocks.append(
            (title, _runs(list(with_text(recs, lambda r: joined(
                r["course"], r["section"], r["room"] or "온라인"))), lambda r: r["prof_id"]))
        )
    if blocks:
        _add_sheet(wb, "교수님별", "교수님별 시간표", blocks, "prof", st,
                   list(PERIODS), PROF_BLOCKS_PER_ROW)

    # ── 3) 학과별 (학년 블록) ───────────────────────────────────────────────
    for dept in sorted({r["department"] for r in records if r["department"]}):
        grades = sorted({r["grade"] for r in records if r["department"] == dept})
        blocks = []
        for g in grades:
            recs = [r for r in records if r["department"] == dept and r["grade"] == g]
            blocks.append(
                (f"{g}학년", _runs(list(with_text(recs, lambda r: joined(
                    r["course"], r["section"], r["professor"], r["room"] or "온라인"))),
                    lambda r: (r["department"], r["grade"])))
            )
        if blocks:
            _add_sheet(wb, dept, f"{dept} 시간표", blocks, "room", st,
                       list(PERIODS), BLOCKS_PER_ROW)

    if not wb.sheetnames:  # nothing scheduled yet — still return a valid file
        wb.create_sheet(title="시간표")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
